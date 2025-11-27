# selenium does not support excel automation that's why we are using openpyxl module to automnate excel


# first import openpyxl modlue 
import openpyxl

#find file path
file_path = "python_data_driven_testing\data.xlsx"

#load workbook
workbook = openpyxl.load_workbook(file_path)

# select sheet
sheet = workbook["Sheet1"]

'''
In an axcel sheet so many rows are available and each row have columns 
first we have to find how many rows and column is available in sheet1
then we will iterate throw row then through column
'''
# finding total rows and column number in excel
rows = sheet.max_row
col = sheet.max_column

#iterating throgh row and evry col of row
# reading all cell data of excel file with below logic
for r in range(1,rows+1):   # here adding one because range last value is excluden means not taken
    for c in range(1,col+1):
        print(sheet.cell(r,c).value,end=" ") 
    print("\n")    


