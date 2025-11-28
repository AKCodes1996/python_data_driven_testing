


import openpyxl

# function for getting toital rows
def get_total_rows(file,sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row

# function for getting total number of column
def get_total_column(file,sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name] 
    return sheet.max_column

# function for writing data
def writing_data(row,col,file,sheet_name,cell_value):
    workbook = openpyxl.load_workbook(file)
    sheet =workbook[sheet_name]
    sheet_name.cell(row,col).value = cell_value
    workbook.save(file)
    
#function for reading specific cell data
def reading_data(file,sheet_name,row,col):
    workbook = openpyxl.load_workbook(file) 
    sheet = workbook[sheet_name]
    return sheet.cell(row,col).value