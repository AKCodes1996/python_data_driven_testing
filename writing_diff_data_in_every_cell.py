# below code is for writing different data in every rows and colmn



import openpyxl

#(user can give any file path )
file = "python_data_driven_testing\data3.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]


# as per below below logic we have to manually write specific data for specific cell

#writing 1st row data for evry column
sheet.cell(1,1).value = "name"
sheet.cell(1,2).value = "age"


#writing 2nd row data for evry column
sheet.cell(2,1).value = "AMIT"
sheet.cell(2,2).value = 20


#writing 3rd row data for evry column
sheet.cell(3,1).value = "kittu"
sheet.cell(3,2).value = 18

workbook.save(file)